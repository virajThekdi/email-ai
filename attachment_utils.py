from __future__ import annotations

import csv
import io
from email.message import Message
from zipfile import BadZipFile


MAX_ATTACHMENT_CHARS = 12000


def extract_attachments_from_message(message: Message) -> tuple[list[str], str]:
    names: list[str] = []
    chunks: list[str] = []
    if not message.is_multipart():
        return names, ""

    for part in message.walk():
        filename = part.get_filename()
        if not filename:
            continue
        payload = part.get_payload(decode=True)
        if not payload:
            continue
        names.append(filename)
        text = extract_attachment_text(filename, payload)
        if text:
            chunks.append(f"Attachment: {filename}\n{text}")

    combined = "\n\n".join(chunks)
    return names, combined[:MAX_ATTACHMENT_CHARS]


def extract_attachment_text(filename: str, payload: bytes) -> str:
    lower = filename.lower()
    try:
        if lower.endswith(".pdf"):
            return _extract_pdf(payload)
        if lower.endswith(".xlsx"):
            return _extract_xlsx(payload)
        if lower.endswith(".csv"):
            return _extract_csv(payload)
        if lower.endswith(".docx"):
            return _extract_docx(payload)
        if lower.endswith((".txt", ".md", ".log")):
            return payload.decode("utf-8", errors="replace")[:MAX_ATTACHMENT_CHARS]
    except Exception as exc:
        return f"Could not read attachment text: {exc}"
    return ""


def _extract_pdf(payload: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(payload))
    pages = []
    for page in reader.pages[:8]:
        pages.append(page.extract_text() or "")
    return "\n".join(pages)[:MAX_ATTACHMENT_CHARS]


def _extract_xlsx(payload: bytes) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    rows = []
    for sheet in workbook.worksheets[:4]:
        rows.append(f"Sheet: {sheet.title}")
        for row in sheet.iter_rows(max_row=40, max_col=12, values_only=True):
            values = [str(value) for value in row if value not in (None, "")]
            if values:
                rows.append(" | ".join(values))
    return "\n".join(rows)[:MAX_ATTACHMENT_CHARS]


def _extract_csv(payload: bytes) -> str:
    text = payload.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = []
    for index, row in enumerate(reader):
        if index >= 80:
            break
        rows.append(" | ".join(cell for cell in row if cell))
    return "\n".join(rows)[:MAX_ATTACHMENT_CHARS]


def _extract_docx(payload: bytes) -> str:
    from docx import Document

    try:
        document = Document(io.BytesIO(payload))
    except BadZipFile:
        return ""
    return "\n".join(paragraph.text for paragraph in document.paragraphs)[:MAX_ATTACHMENT_CHARS]
